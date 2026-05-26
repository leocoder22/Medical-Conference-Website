from io import BytesIO
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import render_template, flash, redirect, url_for, current_app, request
from flask_login import login_required, current_user
from sqlalchemy import func

from . import admin_bp
from app.extensions import db
from app.models import (
    User,
    ConferenceRegistration,
    AbstractSubmission,
    Speaker,
    ScheduleItem,
)
from app.utils.email_utils import (
    send_payment_verified_email,
    send_abstract_accepted_email,
    send_abstract_rejected_email
)


def is_admin():
    admin_email = current_app.config.get("ADMIN_EMAIL")
    return current_user.is_authenticated and current_user.email == admin_email


def admin_required_redirect():
    if not is_admin():
        flash("Access denied. Admins only.", "danger")
        return redirect(url_for("user.abstract_portal"))
    return None


@admin_bp.route("/dashboard")
@login_required
def dashboard():
    denied = admin_required_redirect()
    if denied:
        return denied

    users = User.query.order_by(User.id.desc()).all()
    registrations = ConferenceRegistration.query.order_by(ConferenceRegistration.created_at.desc()).all()
    abstracts = AbstractSubmission.query.order_by(AbstractSubmission.created_at.desc()).all()
    speakers = Speaker.query.order_by(Speaker.id.asc()).all()
    schedule_items = ScheduleItem.query.order_by(
        ScheduleItem.day_label.asc(),
        ScheduleItem.display_order.asc(),
        ScheduleItem.id.asc()
    ).all()

    return render_template(
        "admin/dashboard.html",
        users=users,
        registrations=registrations,
        abstracts=abstracts,
        speakers=speakers,
        schedule_items=schedule_items
    )

@admin_bp.route("/export/registrations-abstracts")
@login_required
def export_registrations_abstracts():
    denied = admin_required_redirect()
    if denied:
        return denied

    wb = Workbook()
    ws = wb.active
    ws.title = "Registrations_Abstracts"

    headers = [
        "Registration ID",
        "Full Name",
        "Age",
        "Gender",
        "Email",
        "Phone",
        "Institution",
        "Designation",
        "Department",
        "Council / Registration Authority",
        "Medical Council Registration Number",
        "Registration Type",
        "Registration Fee",
        "Transaction ID",
        "Payment Verified",
        "Registration Date",
        "Abstract DB ID",
        "Abstract Code",
        "Abstract Title",
        "Abstract Category",
        "Authors",
        "Presenting Author",
        "Presentation Preference",
        "Abstract Status",
        "Accepted As",
        "Abstract Submitted Date",
    ]

    ws.append(headers)

    registrations = ConferenceRegistration.query.order_by(
        ConferenceRegistration.created_at.desc()
    ).all()

    for reg in registrations:
        user_abstracts = reg.user.abstracts if reg.user else []

        if user_abstracts:
            for abstract in user_abstracts:
                ws.append([
                    reg.id,
                    reg.full_name,
                    reg.age,
                    reg.gender,
                    reg.email,
                    reg.phone,
                    reg.institution_name,
                    reg.designation,
                    reg.department,
                    reg.council_name,
                    reg.medical_council_registration_number,
                    reg.registration_type,
                    reg.registration_fee,
                    reg.transaction_id,
                    "Verified" if reg.payment_verified else "Pending",
                    reg.created_at.strftime("%d-%m-%Y %H:%M") if reg.created_at else "",
                    abstract.id,
                    abstract.abstract_code if abstract.abstract_code else "",
                    abstract.title,
                    abstract.category,
                    abstract.authors,
                    abstract.presenting_author,
                    abstract.presentation_preference,
                    abstract.status,
                    abstract.accepted_as if abstract.accepted_as else "",
                    abstract.created_at.strftime("%d-%m-%Y %H:%M") if abstract.created_at else "",
                ])
        else:
            ws.append([
                reg.id,
                reg.full_name,
                reg.age,
                reg.gender,
                reg.email,
                reg.phone,
                reg.institution_name,
                reg.designation,
                reg.department,
                reg.council_name,
                reg.medical_council_registration_number,
                reg.registration_type,
                reg.registration_fee,
                reg.transaction_id,
                "Verified" if reg.payment_verified else "Pending",
                reg.created_at.strftime("%d-%m-%Y %H:%M") if reg.created_at else "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ])

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="1E3A8A")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    column_widths = {
        "A": 16,
        "B": 26,
        "C": 8,
        "D": 12,
        "E": 32,
        "F": 16,
        "G": 34,
        "H": 22,
        "I": 22,
        "J": 34,
        "K": 32,
        "L": 26,
        "M": 16,
        "N": 32,
        "O": 18,
        "P": 22,
        "Q": 14,
        "R": 18,
        "S": 42,
        "T": 24,
        "U": 40,
        "V": 24,
        "W": 24,
        "X": 18,
        "Y": 18,
        "Z": 24,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 36
    for row_num in range(2, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 45

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="MAHATROPACON_Registrations_Abstracts.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# -----------------------------
# REGISTRATION / PAYMENT VERIFICATION
# -----------------------------

@admin_bp.route("/registration/<int:registration_id>/verify-payment", methods=["POST"])
@login_required
def verify_registration_payment(registration_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    registration = ConferenceRegistration.query.get_or_404(registration_id)
    registration.payment_verified = True
    db.session.commit()

    email_sent = send_payment_verified_email(
        recipient_email=registration.email,
        participant_name=registration.full_name
    )

    if email_sent:
        flash("Payment marked as verified and confirmation email sent.", "success")
    else:
        flash("Payment marked as verified, but confirmation email could not be sent.", "warning")

    return redirect(url_for("admin.dashboard"))


@admin_bp.route("/registration/<int:registration_id>/unverify-payment", methods=["POST"])
@login_required
def unverify_registration_payment(registration_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    registration = ConferenceRegistration.query.get_or_404(registration_id)
    registration.payment_verified = False
    db.session.commit()

    flash("Payment verification removed.", "warning")
    return redirect(url_for("admin.dashboard"))

# -----------------------------
# ABSTRACT REVIEW CONTROL
# -----------------------------

def generate_abstract_code(abstract, accepted_as):
    prefix = "O" if accepted_as == "Oral" else "P"
    return f"{prefix}{abstract.id:04d}"


@admin_bp.route("/abstract/<int:abstract_id>/accept", methods=["POST"])
@login_required
def accept_abstract(abstract_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    abstract = AbstractSubmission.query.get_or_404(abstract_id)
    accepted_as = request.form.get("accepted_as", "").strip()

    if accepted_as not in ["Oral", "Poster"]:
        flash("Please select Oral or Poster before accepting.", "danger")
        return redirect(url_for("admin.dashboard"))

    abstract.status = "accepted"
    abstract.accepted_as = accepted_as
    abstract.abstract_code = generate_abstract_code(abstract, accepted_as)

    db.session.commit()

    email_sent = send_abstract_accepted_email(
        recipient_email=abstract.user.email,
        participant_name=abstract.user.full_name,
        abstract_title=abstract.title,
        accepted_as=accepted_as,
        abstract_code=abstract.abstract_code
    )

    if email_sent:
        flash("Abstract accepted and email sent.", "success")
    else:
        flash("Abstract accepted, but email could not be sent.", "warning")

    return redirect(url_for("admin.dashboard"))


@admin_bp.route("/abstract/<int:abstract_id>/reject", methods=["POST"])
@login_required
def reject_abstract(abstract_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    abstract = AbstractSubmission.query.get_or_404(abstract_id)

    abstract.status = "rejected"
    abstract.accepted_as = None
    abstract.abstract_code = None

    db.session.commit()

    email_sent = send_abstract_rejected_email(
        recipient_email=abstract.user.email,
        participant_name=abstract.user.full_name,
        abstract_title=abstract.title
    )

    if email_sent:
        flash("Abstract rejected and email sent.", "warning")
    else:
        flash("Abstract rejected, but email could not be sent.", "warning")

    return redirect(url_for("admin.dashboard"))

# -----------------------------
# INLINE SPEAKER MANAGEMENT
# -----------------------------

@admin_bp.route("/speakers/create-inline", methods=["POST"])
@login_required
def create_inline_speaker():
    denied = admin_required_redirect()
    if denied:
        return denied

    speaker = Speaker(
        name="",
        designation="",
        institution="",
        specialization="",
        bio="",
        initials="",
        gradient_class="from-blue-100 to-cyan-100",
        text_color_class="text-blue-700",
        is_featured=False,
        display_order=0
    )

    db.session.add(speaker)
    db.session.commit()

    flash("New speaker card added.", "success")
    return redirect(url_for("main.speakers"))


@admin_bp.route("/speakers/<int:speaker_id>/update-inline", methods=["POST"])
@login_required
def update_inline_speaker(speaker_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    speaker = Speaker.query.get_or_404(speaker_id)

    name = request.form.get("name", "").strip()
    designation = request.form.get("designation", "").strip()
    institution = request.form.get("institution", "").strip()
    bio = request.form.get("bio", "").strip()

    if not name:
        flash("Speaker name is required.", "danger")
        return redirect(url_for("main.speakers"))

    speaker.name = name
    speaker.designation = ""
    speaker.institution = institution
    speaker.specialization = designation
    speaker.bio = bio

    db.session.commit()

    flash("Speaker updated successfully.", "success")
    return redirect(url_for("main.speakers"))


@admin_bp.route("/speakers/<int:speaker_id>/delete-inline", methods=["POST"])
@login_required
def delete_inline_speaker(speaker_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    speaker = Speaker.query.get_or_404(speaker_id)
    db.session.delete(speaker)
    db.session.commit()

    flash("Speaker deleted successfully.", "warning")
    return redirect(url_for("main.speakers"))


# -----------------------------
# SIMPLE INLINE SCHEDULE MANAGEMENT
# -----------------------------

@admin_bp.route("/schedule/create-day", methods=["POST"])
@login_required
def create_schedule_day():
    denied = admin_required_redirect()
    if denied:
        return denied

    existing_days = db.session.query(ScheduleItem.day_label).distinct().all()
    existing_day_names = [d[0] for d in existing_days]

    next_number = 1
    while f"Day {next_number}" in existing_day_names:
        next_number += 1

    new_day_label = f"Day {next_number}"

    item = ScheduleItem(
        day_label=new_day_label,
        section_title=new_day_label,
        from_time="",
        to_time="",
        title="",
        speaker="",
        venue="",
        display_order=1
    )

    db.session.add(item)
    db.session.commit()

    flash(f"{new_day_label} added.", "success")
    return redirect(url_for("main.schedule"))


@admin_bp.route("/schedule/create-inline-session", methods=["POST"])
@login_required
def create_inline_schedule_session():
    denied = admin_required_redirect()
    if denied:
        return denied

    day_label = request.form.get("day_label", "").strip()

    if not day_label:
        flash("Day is required.", "danger")
        return redirect(url_for("main.schedule"))

    next_order = (
        db.session.query(func.max(ScheduleItem.display_order))
        .filter(ScheduleItem.day_label == day_label)
        .scalar() or 0
    ) + 1

    item = ScheduleItem(
        day_label=day_label,
        section_title=day_label,
        from_time="",
        to_time="",
        title="",
        speaker="",
        venue="",
        display_order=next_order
    )

    db.session.add(item)
    db.session.commit()

    flash("Session added.", "success")
    return redirect(url_for("main.schedule"))


@admin_bp.route("/schedule/day-card/update", methods=["POST"])
@login_required
def update_schedule_day_card():
    denied = admin_required_redirect()
    if denied:
        return denied

    original_day_label = request.form.get("original_day_label", "").strip()
    new_day_title = request.form.get("day_title", "").strip()

    if not original_day_label or not new_day_title:
        flash("Day title is required.", "danger")
        return redirect(url_for("main.schedule"))

    items = ScheduleItem.query.filter_by(day_label=original_day_label).all()

    session_ids = request.form.getlist("session_id[]")
    from_times = request.form.getlist("from_time[]")
    to_times = request.form.getlist("to_time[]")
    titles = request.form.getlist("title[]")
    speakers = request.form.getlist("speaker[]")
    venues = request.form.getlist("venue[]")

    for idx, session_id in enumerate(session_ids):
        try:
            item = ScheduleItem.query.get(int(session_id))
        except (TypeError, ValueError):
            item = None

        if item:
            item.day_label = new_day_title
            item.section_title = new_day_title
            item.from_time = from_times[idx].strip() if idx < len(from_times) else ""
            item.to_time = to_times[idx].strip() if idx < len(to_times) else ""
            item.title = titles[idx].strip() if idx < len(titles) else ""
            item.speaker = speakers[idx].strip() if idx < len(speakers) else ""
            item.venue = venues[idx].strip() if idx < len(venues) else ""
            item.display_order = idx + 1

    db.session.commit()

    flash("Schedule updated.", "success")
    return redirect(url_for("main.schedule"))


@admin_bp.route("/schedule/day-card/save-and-add-session", methods=["POST"])
@login_required
def save_day_and_add_session():
    denied = admin_required_redirect()
    if denied:
        return denied

    original_day_label = request.form.get("original_day_label", "").strip()
    new_day_title = request.form.get("day_title", "").strip()

    if not original_day_label or not new_day_title:
        flash("Day title is required.", "danger")
        return redirect(url_for("main.schedule"))

    items = ScheduleItem.query.filter_by(day_label=original_day_label).all()

    session_ids = request.form.getlist("session_id[]")
    from_times = request.form.getlist("from_time[]")
    to_times = request.form.getlist("to_time[]")
    titles = request.form.getlist("title[]")
    speakers = request.form.getlist("speaker[]")
    venues = request.form.getlist("venue[]")
    

    for idx, session_id in enumerate(session_ids):
        try:
            item = ScheduleItem.query.get(int(session_id))
        except (TypeError, ValueError):
            item = None

        if item:
            item.day_label = new_day_title
            item.section_title = new_day_title
            item.from_time = from_times[idx].strip() if idx < len(from_times) else ""
            item.to_time = to_times[idx].strip() if idx < len(to_times) else ""
            item.title = titles[idx].strip() if idx < len(titles) else ""
            item.speaker = speakers[idx].strip() if idx < len(speakers) else ""
            item.venue = venues[idx].strip() if idx < len(venues) else ""
            item.display_order = idx + 1

    db.session.flush()

    next_order = (
        db.session.query(func.max(ScheduleItem.display_order))
        .filter(ScheduleItem.day_label == new_day_title)
        .scalar() or 0
    ) + 1

    new_item = ScheduleItem(
        day_label=new_day_title,
        section_title=new_day_title,
        from_time="",
        to_time="",
        title="",
        speaker="",
        venue="",
        display_order=next_order
    )

    db.session.add(new_item)
    db.session.commit()

    flash("Saved and new session added.", "success")
    return redirect(url_for("main.schedule"))


@admin_bp.route("/schedule/<int:item_id>/delete-inline", methods=["POST"])
@login_required
def delete_inline_schedule_item(item_id):
    denied = admin_required_redirect()
    if denied:
        return denied

    item = ScheduleItem.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()

    flash("Session deleted.", "warning")
    return redirect(url_for("main.schedule"))

@admin_bp.route("/schedule/day/move", methods=["POST"])
@login_required
def move_schedule_day():
    denied = admin_required_redirect()
    if denied:
        return denied

    day_label = request.form.get("day_label", "").strip()
    direction = request.form.get("direction", "").strip()

    if not day_label or direction not in ["up", "down"]:
        flash("Invalid move request.", "danger")
        return redirect(url_for("main.schedule"))

    day_labels = [
        row[0] for row in db.session.query(ScheduleItem.day_label)
        .distinct()
        .order_by(ScheduleItem.day_label.asc())
        .all()
    ]

    if day_label not in day_labels:
        flash("Day not found.", "warning")
        return redirect(url_for("main.schedule"))

    index = day_labels.index(day_label)

    if direction == "up" and index == 0:
        return redirect(url_for("main.schedule"))

    if direction == "down" and index == len(day_labels) - 1:
        return redirect(url_for("main.schedule"))

    swap_index = index - 1 if direction == "up" else index + 1
    swap_day_label = day_labels[swap_index]

    temp_label = f"__temp_day_{day_label}__"

    current_items = ScheduleItem.query.filter_by(day_label=day_label).all()
    swap_items = ScheduleItem.query.filter_by(day_label=swap_day_label).all()

    for item in current_items:
        item.day_label = temp_label

    db.session.flush()

    for item in swap_items:
        item.day_label = day_label

    db.session.flush()

    for item in current_items:
        item.day_label = swap_day_label

    db.session.commit()

    flash("Day position updated.", "success")
    return redirect(url_for("main.schedule"))